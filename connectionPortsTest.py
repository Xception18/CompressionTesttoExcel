import serial
import serial.tools.list_ports
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Konfigurasi default untuk koneksi serial
BAUD_RATE = 9600
DATA = 8

def list_com_ports():
    """
    Menampilkan dan mengembalikan daftar COM ports yang tersedia
    """
    ports = serial.tools.list_ports.comports()
    
    if not ports:
        print("No COM ports found")
        return []
    
    print("Available COM ports:")
    available_ports = []
    
    for port in ports:
        print(f"Port: {port.device}")
        print(f"Description: {port.description}")
        print(f"Hardware ID: {port.hwid}")
        print("-" * 40)
        available_ports.append(port.device)
    
    return available_ports

def koneksi(port, baud=BAUD_RATE, data=DATA):
    """
    Menguji koneksi ke port serial yang ditentukan
    
    Args:
        port (str): Nama port (misal: 'COM3' atau '/dev/ttyUSB0')
        baud (int): Baud rate (default: 9600)
        data (int): Data bits (default: 8)
    
    Returns:
        str: Status koneksi
    """
    try:
        ser = serial.Serial(port, baud, data, timeout=None)
        if ser.is_open:
            print(f"Berhasil terhubung ke {port}")
            ser.close()
            return "Tersambung !"
        else:
            return "Gagal membuka port"
    except Exception as e:
        pesanError = str(e)
        print(f"Error connecting to {port}: {pesanError}")
        return f"Error: {pesanError}"

def baca_dan_simpan_ke_excel(port, baud=BAUD_RATE, data=DATA, excel_file="pressure_data.xlsx"):
    """
    Membaca nilai tekanan dari perangkat serial secara terus-menerus
    dan menyimpan ke Excel per baris tanpa timeout
    
    Args:
        port (str): Nama port (misal: 'COM3' atau '/dev/ttyUSB0')
        baud (int): Baud rate (default: 9600)
        data (int): Data bits (default: 8)
        excel_file (str): Nama file Excel untuk menyimpan data
    """
    try:
        # Buka koneksi serial tanpa timeout
        ser = serial.Serial(port, baud, data, timeout=None)
        
        if not ser.is_open:
            print("Gagal membuka port")
            return
        
        print(f"Terhubung ke {port}")
        print(f"Baud rate: {baud}")
        print(f"Data akan disimpan ke: {excel_file}")
        print("=" * 60)
        
        # Buat atau load Excel workbook
        if os.path.exists(excel_file):
            wb = load_workbook(excel_file)
            ws = wb.active
            print(f"File Excel '{excel_file}' ditemukan, melanjutkan data...")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Pressure Data"
            # Header
            ws.append(["No", "Timestamp", "Nilai KN", "Raw Data"])
            wb.save(excel_file)
            print(f"File Excel '{excel_file}' dibuat baru")
        
        print("\nMulai membaca data...")
        print("Tekan Ctrl+C untuk menghentikan\n")
        
        counter = ws.max_row if ws.max_row > 1 else 1
        
        while True:
            try:
                # Baca data dari serial (blocking, tanpa timeout)
                if datanya := ser.readline():
                    data_str = datanya.decode("utf-8", errors="ignore").strip()
                    
                    if data_str:  # Jika ada data
                        print(f"[{datetime.now().strftime('%H:%M:%S')}] Raw: {data_str}")
                        
                        # Cek apakah data mengandung "ovalue"
                        if "ovalue" in data_str.lower():
                            try:
                                # Extract nilai KN
                                parts = data_str.split()
                                if len(parts) >= 2:
                                    nilaiKN = parts[1]
                                    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                    
                                    # Simpan ke Excel
                                    ws.append([counter, timestamp, nilaiKN, data_str])
                                    wb.save(excel_file)
                                    
                                    print(f"✓ Data #{counter} tersimpan: {nilaiKN} KN")
                                    print("-" * 60)
                                    counter += 1
                                else:
                                    print("⚠ Format data tidak lengkap, menunggu data berikutnya...")
                                    
                            except Exception as e:
                                print(f"⚠ Error parsing data: {e}")
                                continue
                        
            except KeyboardInterrupt:
                print("\n\nProses dihentikan oleh user")
                break
            except Exception as e:
                print(f"⚠ Error membaca data: {e}")
                continue
        
        # Tutup koneksi
        ser.close()
        wb.close()
        print(f"\n✓ Koneksi ditutup")
        print(f"✓ Total data tersimpan: {counter - 1}")
        print(f"✓ File Excel: {excel_file}")
        
    except Exception as e:
        pesanError = str(e)
        print(f"Error: {pesanError}")

def pilih_port_dan_mulai_logging():
    """
    Memungkinkan pengguna memilih port untuk logging ke Excel
    """
    ports = list_com_ports()
    
    if not ports:
        return
    
    print(f"\nDitemukan {len(ports)} COM ports")
    print("Pilih port untuk logging data:")
    
    for i, port in enumerate(ports, 1):
        print(f"{i}. {port}")
    
    try:
        choice = int(input("\nMasukkan nomor port (atau 0 untuk keluar): "))
        
        if choice == 0:
            print("Keluar...")
            return
        
        if 1 <= choice <= len(ports):
            selected_port = ports[choice - 1]
            
            # Input nama file Excel
            excel_file = input("\nNama file Excel (tekan Enter untuk 'pressure_data.xlsx'): ").strip()
            if not excel_file:
                excel_file = "pressure_data.xlsx"
            
            if not excel_file.endswith('.xlsx'):
                excel_file += '.xlsx'
            
            # Input baud rate (opsional)
            baud_input = input(f"Baud rate (tekan Enter untuk {BAUD_RATE}): ").strip()
            baud = int(baud_input) if baud_input else BAUD_RATE
            
            print(f"\n{'='*60}")
            print(f"Port: {selected_port}")
            print(f"Baud rate: {baud}")
            print(f"File Excel: {excel_file}")
            print(f"{'='*60}\n")
            
            # Mulai logging
            baca_dan_simpan_ke_excel(selected_port, baud, DATA, excel_file)
        else:
            print("Pilihan tidak valid!")
            
    except ValueError:
        print("Masukkan nomor yang valid!")
    except KeyboardInterrupt:
        print("\nOperasi dibatalkan oleh user.")

# Main program
if __name__ == "__main__":
    print("=" * 60)
    print("Serial Port Pressure Logger to Excel")
    print("=" * 60)
    
    while True:
        print("\nMenu:")
        print("1. Lihat daftar COM ports")
        print("2. Mulai logging data ke Excel")
        print("3. Keluar")
        
        try:
            choice = input("\nPilih opsi (1-3): ").strip()
            
            if choice == '1':
                print()
                ports = list_com_ports()
                print(f"\nDitemukan {len(ports)} COM ports")
                
            elif choice == '2':
                pilih_port_dan_mulai_logging()
                
            elif choice == '3':
                print("Terima kasih!")
                break
                
            else:
                print("Opsi tidak valid! Pilih 1-3.")
                
        except KeyboardInterrupt:
            print("\n\nProgram dihentikan. Terima kasih!")
            break
        except Exception as e:
            print(f"Terjadi error: {e}")
